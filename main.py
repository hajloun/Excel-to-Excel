import PySimpleGUI as sg # type: ignore
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string


# LAYOUT
sg.theme('SystemDefault1')
layout = [
    [sg.Text("Zadejte ID pro vyhledání:", size=(22, 1)), sg.InputText(key='-ID-')],
    [sg.Text("Zdrojový Excel soubor:", size=(22, 1)), 
     sg.Input(key='-SOURCE_FILE-'), 
     sg.FileBrowse("Vybrat...", file_types=(("Excel Files", "*.xlsx;*.xls"),))],
    [sg.Text("Cílový Excel soubor:", size=(22, 1)), 
    sg.Input(key='-TARGET_FILE-', enable_events=True),
    sg.FileBrowse("Vybrat...", file_types=(("Excel Files", "*.xlsx;*.xls"),))],
    [sg.Text("Vyberte cílový list:", size=(22, 1)), 
    sg.Combo([], key='-SHEET_NAME-', size=(40, 1), readonly=True)],
    [sg.Button("Zpracovat data", key='-PROCESS-'), sg.Button("Konec")],
    [sg.Text("Průběh a informace:")],
    [sg.Multiline(size=(85, 15), key='-OUTPUT-', disabled=True, autoscroll=True, reroute_stdout=True)]
]
window = sg.Window("Excel to Excel", layout)


# EVENT LOOP 
while True:
    event, values = window.read()
    if event == '-TARGET_FILE-':
        target_path = values['-TARGET_FILE-']
        if target_path:
            try:
                xls = pd.ExcelFile(target_path)
                sheet_names = xls.sheet_names
                window['-SHEET_NAME-'].update(values=sheet_names, set_to_index=0)
            except Exception as e:
                window['-SHEET_NAME-'].update(values=[], value='')
                print(f"CHYBA při čtení listů: {e}")
    if event == sg.WIN_CLOSED or event == 'Konec':
        break
    if event == '-PROCESS-':
        id_to_find = values['-ID-']
        source_path = values['-SOURCE_FILE-']
        target_path = values['-TARGET_FILE-']
        if not id_to_find or not source_path or not target_path:
            print("CHYBA: Všechna pole musí být vyplněna.")
            continue
        print(f" Hledané ID: {id_to_find}")
        print(f" Zdrojový soubor: {source_path}")
        print(f" Cílový soubor: {target_path}")
        print("-" * 30)
    

        # LOAD FILE
        df_source = None
        try:
            df_source = pd.read_excel(source_path, engine="openpyxl", header=9)

        except FileNotFoundError:
            print(f"CHYBA: Zdrojový soubor nebyl nalezen na cestě: {source_path}")
            continue
        except Exception as e:
            print(f"CHYBA: Nepodařilo se načíst zdrojový soubor.")
            print(f"Podrobnosti chyby: {e}")


        # FIND ROW
        SOURCE_ID_COLUMN_NAME = 'ID' # CHANGE TO EXCEL ID COLUMN!!
        COLUMN_MAPPING = {
    'ID': 'F42 ID',
    'Name': 'Function',
    'Name (English)': 'Function (English)',
    'Implementation managers': 'FuReV',
    'FuLi contact person': 'FuReV support',
    'Einsatz zu': 'Einsatz zu',
    'Entfall zu': 'Entfall zu',
    'Funktionscluster (VW) / Solution (CARIAD)': 'Cluster'
}
        found_row = None
        try:
            id_to_find_str = str(id_to_find)
            source_id_column_str = df_source[SOURCE_ID_COLUMN_NAME].astype(str)
            found_data = df_source[source_id_column_str == id_to_find_str]
            if found_data.empty:
                print(f"CHYBA: ID '{id_to_find}' nebylo ve zdrojovém souboru nalezeno.")
                print("-" * 30)
                continue
            found_row = found_data.iloc[0]
            print("Záznam s daným ID byl úspěšně nalezen.")

        except KeyError:
            print(f"CHYBA: V souboru chybí sloupec s názvem '{SOURCE_ID_COLUMN_NAME}'.")
            continue
        except Exception as e:
            print(f"CHYBA: Vyskytla se chyba při hledání dat: {e}")
            continue
        

        # PREPARE ROW AND SAVE WITH OPENPYXL, EXTEND TABLE
        try:
            workbook = load_workbook(target_path)
            sheet = workbook[selected_sheet_name]
        if not sheet.tables:
                print("CHYBA: V cílovém listu nebyl nalezen žádný formátovaný objekt Tabulka.")
                continue
        table_name = list(sheet.tables.keys())[0]
        table = sheet.tables[table_name]
        start_cell, end_cell = table.ref.split(':')
        end_col_idx = column_index_from_string(end_cell.rstrip('0123456789'))
        last_table_row = int(''.join(filter(str.isdigit, end_cell)))
        sheet.insert_rows(last_table_row)
        print(f"Prázdný řádek byl vložen na pozici {last_table_row}.")
        new_row_dict = {}
        for source_col, target_col in COLUMN_MAPPING.items():
                if source_col in found_row:
                    new_row_dict[target_col] = found_row[source_col]
        target_columns_order = [cell.value for cell in sheet[10]]
        for idx, col_name in enumerate(target_columns_order, 1):
                if col_name in new_row_dict:
                    sheet.cell(row=last_table_row, column=idx).value = new_row_dict[col_name]
        table.ref = f"{start_cell}:{get_column_letter(end_col_idx)}{last_table_row + 1}"
        workbook.save(target_path)
        print("\nHotovo! Řádek byl vložen do tabulky.")

        except Exception as e:
            print(f"CHYBA: Vyskytla se chyba při zpracování cílového souboru: {e}")
            if 'PermissionError' in str(e):
                print("Není soubor otevřený v Excelu?")
            continue
        

        print("-" * 30)

window.close()
